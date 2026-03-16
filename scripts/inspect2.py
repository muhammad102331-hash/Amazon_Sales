import pandas as pd
import openpyxl
wb = openpyxl.load_workbook('amazon_sales.xlsx', read_only=False, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb.active
print("Dims:", ws.dimensions)
print("Max row:", ws.max_row, "Max col:", ws.max_column)
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("Headers:", headers)
